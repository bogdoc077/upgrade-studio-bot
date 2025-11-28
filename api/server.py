from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBasic, HTTPBasicCredentials, HTTPBearer
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any, Optional
import secrets
from datetime import datetime, timedelta
import bcrypt
import jwt
from jwt import InvalidTokenError
import sys
from pathlib import Path
from pydantic import BaseModel
import os
import uuid
import shutil
from PIL import Image
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Додаємо шлях до проєкту
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from database.models import get_database
from database.encryption import settings_manager
from config import settings

# Pydantic моделі
class LoginRequest(BaseModel):
    username: str
    password: str

class AdminCreate(BaseModel):
    username: str
    email: str
    password: str
    first_name: str
    last_name: Optional[str] = None
    is_active: bool = True

class AdminUpdate(BaseModel):
    email: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    is_active: Optional[bool] = None

class PasswordChange(BaseModel):
    current_password: str

class SettingUpdate(BaseModel):
    key: str
    value: str
    value_type: str = "string"
    category: str = "general"
    is_sensitive: bool = False
    description: Optional[str] = None


app = FastAPI(title="Upgrade Studio Bot API", version="1.0.0")

# Створюємо папки для завантажень якщо не існують
UPLOAD_DIR = PROJECT_ROOT / "uploads"
BROADCASTS_DIR = UPLOAD_DIR / "broadcasts"
BROADCASTS_DIR.mkdir(parents=True, exist_ok=True)

# Монтуємо статичні файли
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

# CORS для Next.js
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "https://admin.upgrade21.com"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Health check endpoint
@app.get("/health")
async def health_check():
    """Health check endpoint для моніторингу"""
    try:
        # Перевіримо з'єднання з базою даних
        db = get_database()
        cursor = db.cursor()
        cursor.execute("SELECT 1")
        cursor.close()
        db.close()
        
        return {
            "status": "healthy",
            "timestamp": datetime.utcnow().isoformat(),
            "version": "1.0.0",
            "services": {
                "database": "connected",
                "api": "running"
            }
        }
    except Exception as e:
        return {
            "status": "unhealthy",
            "timestamp": datetime.utcnow().isoformat(),
            "version": "1.0.0",
            "error": str(e),
            "services": {
                "database": "error",
                "api": "running"
            }
        }

# Basic Authentication
# JWT налаштування
JWT_SECRET = settings_manager.get("jwt_secret", "your-secret-key-change-in-production")
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION = timedelta(hours=24)

security = HTTPBasic(auto_error=False)
bearer_security = HTTPBearer(auto_error=False)

def create_access_token(data: dict) -> str:
    """Створити JWT токен"""
    to_encode = data.copy()
    expire = datetime.utcnow() + JWT_EXPIRATION
    to_encode.update({"exp": expire})
    # Переконуємось що sub є рядком
    if "sub" in to_encode:
        to_encode["sub"] = str(to_encode["sub"])
    encoded_jwt = jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return encoded_jwt

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Перевірити пароль"""
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

def hash_password(password: str) -> str:
    """Захешувати пароль"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def get_admin_by_username(username: str) -> Optional[Dict]:
    """Отримати адміна за ім'ям користувача"""
    db = get_database()
    cursor = db.cursor(dictionary=True)
    
    try:
        cursor.execute("SELECT * FROM admins WHERE username = %s AND is_active = TRUE", (username,))
        admin = cursor.fetchone()
        return admin
    finally:
        cursor.close()
        db.close()

def get_admin_by_id(admin_id: int) -> Optional[Dict]:
    """Отримати адміна за ID"""
    db = get_database()
    cursor = db.cursor(dictionary=True)
    
    try:
        cursor.execute("SELECT * FROM admins WHERE id = %s AND is_active = TRUE", (admin_id,))
        admin = cursor.fetchone()
        return admin
    finally:
        cursor.close()
        db.close()

def get_current_admin_from_token(token: str = Depends(bearer_security)) -> Dict:
    """Отримати поточного адміна з JWT токена"""
    if not token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token required",
        )
    
    try:
        payload = jwt.decode(token.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        admin_id_str: str = payload.get("sub")
        if admin_id_str is None:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid token",
            )
        admin_id = int(admin_id_str)  # Конвертуємо назад в int
    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token expired",
        )
    except (jwt.InvalidTokenError, ValueError):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid token",
        )
    
    admin = get_admin_by_id(admin_id)
    if admin is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Admin not found",
        )
    
    return admin

def get_current_admin(credentials: Optional[HTTPBasicCredentials] = Depends(security)):
    """Перевірка базової авторизації адміна (для зворотної сумісності)"""
    if not credentials:
        raise HTTPException(
            status_code=401,
            detail="Basic Auth credentials required",
            headers={"WWW-Authenticate": "Basic"},
        )
    
    # Спочатку перевіряємо в базі даних
    admin = get_admin_by_username(credentials.username)
    if admin and verify_password(credentials.password, admin["password_hash"]):
        return admin
    
    # Якщо не знайдено в БД, перевіряємо старі дефолтні креди
    correct_username = secrets.compare_digest(credentials.username, "admin")
    correct_password = secrets.compare_digest(credentials.password, "Qwerty21")
    
    if correct_username and correct_password:
        return {"username": "admin", "id": 0, "is_superadmin": True, "can_manage_users": True, "can_manage_payments": True, "can_manage_settings": True, "can_manage_admins": True}
    
    raise HTTPException(
        status_code=401,
        detail="Invalid credentials",
        headers={"WWW-Authenticate": "Basic"},
    )

def get_current_admin_flexible(
    credentials: Optional[HTTPBasicCredentials] = Depends(security),
    token: Optional[HTTPBearer] = Depends(bearer_security)
) -> Dict:
    """Гнучка авторизація - підтримує як Basic Auth так і JWT"""
    # Спочатку пробуємо JWT токен
    if token and hasattr(token, 'credentials') and token.credentials:
        try:
            payload = jwt.decode(token.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            admin_id_str: str = payload.get("sub")
            if admin_id_str:
                admin_id = int(admin_id_str)  # Конвертуємо назад в int
                admin = get_admin_by_id(admin_id)
                if admin:
                    return admin
        except (jwt.ExpiredSignatureError, jwt.InvalidTokenError, ValueError):
            pass
    
    # Якщо JWT не працює, пробуємо Basic Auth
    if credentials:
        try:
            return get_current_admin(credentials)
        except HTTPException:
            pass
    
    raise HTTPException(
        status_code=401,
        detail="Not authenticated",
    )

def check_admin_permission(admin: Dict, permission: str) -> bool:
    """Перевірити дозвіл адміна - всі адміни мають повний доступ"""
    return True

@app.get("/api/dashboard")
async def get_dashboard(admin: Dict = Depends(get_current_admin_flexible)) -> Dict[str, Any]:
    """Отримати статистику для дашборду"""
    try:
        db = get_database()
        cursor = db.cursor(dictionary=True)
        
        # Total users
        cursor.execute("SELECT COUNT(*) as total FROM users")
        result = cursor.fetchone()
        total_users = result["total"] if result else 0
        
        # Active users (subscription_active = 1)
        cursor.execute("SELECT COUNT(*) as active FROM users WHERE subscription_active = 1")
        result = cursor.fetchone()
        active_users = result["active"] if result else 0
        
        # Inactive users (subscription_active = 0)
        cursor.execute("SELECT COUNT(*) as inactive FROM users WHERE subscription_active = 0")
        result = cursor.fetchone()
        inactive_users = result["inactive"] if result else 0
        
        # Total revenue from successful payments (сума вже в євро)
        cursor.execute("SELECT COALESCE(SUM(amount), 0) as total FROM payments WHERE status IN ('succeeded', 'completed')")
        result = cursor.fetchone()
        if result and result["total"]:
            from decimal import Decimal
            total_amount = result["total"]
            if isinstance(total_amount, Decimal):
                total_revenue = float(total_amount)  # Сума вже в євро
            else:
                total_revenue = float(total_amount)
        else:
            total_revenue = 0.0
        
        # Today's payments count
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM payments 
            WHERE status IN ('succeeded', 'completed') 
            AND DATE(created_at) = CURDATE()
        """)
        result = cursor.fetchone()
        payments_today = result["count"] if result else 0
        
        cursor.close()
        db.close()
        
        return {
            "total_users": total_users,
            "active_users": active_users,
            "inactive_users": inactive_users,
            "total_revenue": total_revenue,  # Вже в євро
            "payments_today": payments_today
        }
    except Exception as e:
        if 'db' in locals():
            try:
                db.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/api/users")
async def get_users(
    page: int = 1, 
    limit: int = 50, 
    search: str = "", 
    admin: Dict = Depends(get_current_admin_flexible)
) -> Dict[str, Any]:
    """Отримати список користувачів"""
    try:
        db = get_database()
        cursor = db.cursor(dictionary=True)
        
        offset = (page - 1) * limit
        
        if search:
            search_query = f"%{search}%"
            count_sql = """
                SELECT COUNT(*) as total 
                FROM users 
                WHERE CAST(telegram_id AS CHAR) LIKE %s OR first_name LIKE %s OR last_name LIKE %s
            """
            cursor.execute(count_sql, (search_query, search_query, search_query))
            result = cursor.fetchone()
            total_users = result["total"] if result else 0
            
            users_sql = """
                SELECT * FROM users 
                WHERE CAST(telegram_id AS CHAR) LIKE %s OR first_name LIKE %s OR last_name LIKE %s
                ORDER BY created_at DESC 
                LIMIT %s OFFSET %s
            """
            cursor.execute(users_sql, (search_query, search_query, search_query, limit, offset))
        else:
            cursor.execute("SELECT COUNT(*) as total FROM users")
            result = cursor.fetchone()
            total_users = result["total"] if result else 0
            
            cursor.execute("SELECT * FROM users ORDER BY created_at DESC LIMIT %s OFFSET %s", (limit, offset))
        
        users = cursor.fetchall() or []
        
        # Convert datetime objects to ISO strings
        for user in users:
            if user.get('created_at'):
                user['created_at'] = user['created_at'].isoformat()
            if user.get('updated_at'):
                user['updated_at'] = user['updated_at'].isoformat()
            if user.get('subscription_end_date'):
                user['subscription_end_date'] = user['subscription_end_date'].isoformat()
            if user.get('next_billing_date'):
                user['next_billing_date'] = user['next_billing_date'].isoformat()
            if user.get('member_since'):
                user['member_since'] = user['member_since'].isoformat()
        
        total_pages = (total_users + limit - 1) // limit if total_users > 0 else 1
        
        cursor.close()
        db.close()
        
        return {
            "data": users,
            "total": total_users,
            "pagination": {
                "current_page": page,
                "total_pages": total_pages,
                "total_users": total_users,
                "per_page": limit
            }
        }
    except Exception as e:
        if 'db' in locals():
            try:
                db.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/api/users/export")
async def export_users(
    admin: Dict = Depends(get_current_admin_flexible)
):
    """Експорт користувачів в Excel"""
    try:
        db = get_database()
        cursor = db.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM users ORDER BY created_at DESC")
        users = cursor.fetchall() or []
        
        # Створюємо Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Користувачі"
        
        # Заголовки
        headers = [
            'ID', 'Telegram ID', 'Username', "Ім'я", 'Прізвище',
            'Статус', 'Підписка активна', 'Підписка до', 
            'Дата реєстрації', 'Останнє оновлення'
        ]
        
        # Стилізація заголовків
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Дані
        for row_num, user in enumerate(users, 2):
            ws.cell(row=row_num, column=1, value=user.get('id'))
            ws.cell(row=row_num, column=2, value=user.get('telegram_id'))
            ws.cell(row=row_num, column=3, value=user.get('username', ''))
            ws.cell(row=row_num, column=4, value=user.get('first_name', ''))
            ws.cell(row=row_num, column=5, value=user.get('last_name', ''))
            ws.cell(row=row_num, column=6, value=user.get('subscription_status', 'inactive'))
            ws.cell(row=row_num, column=7, value='Так' if user.get('subscription_active') == 1 else 'Ні')
            ws.cell(row=row_num, column=8, value=user.get('subscription_end_date', ''))
            ws.cell(row=row_num, column=9, value=user.get('created_at', ''))
            ws.cell(row=row_num, column=10, value=user.get('updated_at', ''))
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Зберігаємо в пам'ять
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        cursor.close()
        db.close()
        
        # Повертаємо файл
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=users_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
        
    except Exception as e:
        if 'db' in locals():
            try:
                db.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Export error: {str(e)}")

@app.get("/api/payments")
async def get_payments(
    page: int = 1, 
    limit: int = 50, 
    admin: Dict = Depends(get_current_admin_flexible)
) -> Dict[str, Any]:
    """Отримати список платежів"""
    try:
        db = get_database()
        cursor = db.cursor(dictionary=True)
        
        offset = (page - 1) * limit
        
        # Count total payments
        cursor.execute("SELECT COUNT(*) as total FROM payments")
        result = cursor.fetchone()
        total_payments = result["total"] if result else 0
        
        # Get payments with user info
        cursor.execute("""
            SELECT p.*, u.telegram_id, u.first_name, u.last_name
            FROM payments p
            JOIN users u ON p.user_id = u.id
            ORDER BY p.created_at DESC
            LIMIT %s OFFSET %s
        """, (limit, offset))
        
        payments = cursor.fetchall() or []
        
        # Convert datetime objects to ISO strings (суми вже в євро)
        for payment in payments:
            if payment.get('created_at'):
                payment['created_at'] = payment['created_at'].isoformat()
            if payment.get('updated_at'):
                payment['updated_at'] = payment['updated_at'].isoformat()
            if payment.get('paid_at'):
                payment['paid_at'] = payment['paid_at'].isoformat()
            # Сума вже в євро, тільки конвертуємо в float
            if payment.get('amount'):
                payment['amount'] = float(payment['amount'])
        
        total_pages = (total_payments + limit - 1) // limit if total_payments > 0 else 1
        
        cursor.close()
        db.close()
        
        return {
            "data": payments,
            "total": total_payments,
            "pagination": {
                "current_page": page,
                "total_pages": total_pages,
                "total_payments": total_payments,
                "per_page": limit
            }
        }
    except Exception as e:
        if 'db' in locals():
            try:
                db.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/api/payments/export")
async def export_payments(
    admin: Dict = Depends(get_current_admin_flexible)
):
    """Експорт платежів в Excel"""
    try:
        db = get_database()
        cursor = db.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT p.*, u.telegram_id, u.first_name, u.last_name, u.username
            FROM payments p
            JOIN users u ON p.user_id = u.id
            ORDER BY p.created_at DESC
        """)
        payments = cursor.fetchall() or []
        
        # Створюємо Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Платежі"
        
        # Заголовки
        headers = [
            'ID', 'User ID', 'Telegram ID', 'Username', "Ім'я користувача",
            'Сума (EUR)', 'Валюта', 'Статус', 'Stripe Invoice ID', 'Stripe Payment ID',
            'Дата створення', 'Дата оплати'
        ]
        
        # Стилізація заголовків
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Дані
        for row_num, payment in enumerate(payments, 2):
            ws.cell(row=row_num, column=1, value=payment.get('id'))
            ws.cell(row=row_num, column=2, value=payment.get('user_id'))
            ws.cell(row=row_num, column=3, value=payment.get('telegram_id'))
            ws.cell(row=row_num, column=4, value=payment.get('username', ''))
            ws.cell(row=row_num, column=5, value=f"{payment.get('first_name', '')} {payment.get('last_name', '')}".strip())
            ws.cell(row=row_num, column=6, value=float(payment.get('amount', 0)))  # Сума вже в євро
            ws.cell(row=row_num, column=7, value=payment.get('currency', '').upper())
            ws.cell(row=row_num, column=8, value=payment.get('status', ''))
            ws.cell(row=row_num, column=9, value=payment.get('stripe_invoice_id', ''))
            ws.cell(row=row_num, column=10, value=payment.get('stripe_payment_intent_id', ''))
            ws.cell(row=row_num, column=11, value=str(payment.get('created_at', '')))
            ws.cell(row=row_num, column=12, value=str(payment.get('paid_at', '')))
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Зберігаємо в пам'ять
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        cursor.close()
        db.close()
        
        # Повертаємо файл
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=payments_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
        
    except Exception as e:
        if 'db' in locals():
            try:
                db.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Export error: {str(e)}")

@app.post("/api/users/{user_id}/subscription")
async def update_user_subscription(
    user_id: int,
    action: str,
    admin: Dict = Depends(get_current_admin_flexible)
):
    """Оновити статус підписки користувача"""
    try:
        if action not in ["activate", "deactivate", "extend"]:
            raise HTTPException(status_code=400, detail="Invalid action")
        
        db = get_database()
        cursor = db.cursor()
        
        if action == "activate":
            cursor.execute("""
                UPDATE users 
                SET subscription_status = 'active',
                    subscription_active = true,
                    subscription_end_date = DATE_ADD(NOW(), INTERVAL 1 MONTH)
                WHERE id = %s
            """, (user_id,))
        elif action == "deactivate":
            cursor.execute("""
                UPDATE users 
                SET subscription_status = 'inactive',
                    subscription_active = false,
                    subscription_end_date = NOW()
                WHERE id = %s
            """, (user_id,))
        elif action == "extend":
            cursor.execute("""
                UPDATE users 
                SET subscription_end_date = DATE_ADD(COALESCE(subscription_end_date, NOW()), INTERVAL 1 MONTH)
                WHERE id = %s
            """, (user_id,))
        
        db.commit()
        cursor.close()
        db.close()
        
        return {"success": True, "message": f"Subscription {action} successful"}
    except Exception as e:
        if 'db' in locals():
            try:
                db.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.delete("/api/users/{user_id}")
async def delete_user(
    user_id: int,
    admin: Dict = Depends(get_current_admin_flexible)
):
    """Видалити користувача та всі пов'язані дані"""
    db = None
    try:
        db = get_database()
        cursor = db.cursor()
        
        # Check if user exists
        cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
        if not cursor.fetchone():
            cursor.close()
            db.close()
            raise HTTPException(status_code=404, detail="User not found")
        
        # Delete from broadcast_queue first (has foreign key to users)
        try:
            cursor.execute("DELETE FROM broadcast_queue WHERE user_id = %s", (user_id,))
            print(f"Deleted broadcast_queue entries for user {user_id}")
        except Exception as e:
            print(f"Broadcast queue error: {str(e)}")
        
        # Delete related payments
        try:
            cursor.execute("DELETE FROM payments WHERE user_id = %s", (user_id,))
            print(f"Deleted payments for user {user_id}")
        except Exception as e:
            print(f"Error deleting payments: {str(e)}")
        
        # Delete reminders (correct table name)
        try:
            cursor.execute("DELETE FROM reminders WHERE user_id = %s", (user_id,))
            print(f"Deleted reminders for user {user_id}")
        except Exception as e:
            print(f"Error deleting reminders: {str(e)}")
        
        # Delete the user
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        print(f"Deleted user {user_id}")
        
        db.commit()
        cursor.close()
        db.close()
        
        return {"success": True, "message": "User deleted successfully"}
    except HTTPException:
        if db:
            try:
                db.rollback()
                db.close()
            except:
                pass
        raise
    except Exception as e:
        print(f"Error deleting user {user_id}: {str(e)}")
        if db:
            try:
                db.rollback()
                db.close()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.post("/api/auth/login")
async def login(login_data: LoginRequest):
    """Авторизація адміна"""
    admin = get_admin_by_username(login_data.username)
    
    if not admin or not verify_password(login_data.password, admin["password_hash"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password"
        )
    
    # Оновлюємо час останнього входу
    db = get_database()
    cursor = db.cursor()
    try:
        cursor.execute(
            "UPDATE admins SET last_login_at = NOW() WHERE id = %s",
            (admin["id"],)
        )
        db.commit()
    finally:
        cursor.close()
        db.close()
    
    # Створюємо JWT токен
    access_token = create_access_token(data={"sub": admin["id"]})
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "admin": {
            "id": admin["id"],
            "username": admin["username"],
            "email": admin["email"],
            "first_name": admin["first_name"],
            "last_name": admin["last_name"],
            "is_superadmin": admin["is_superadmin"],
        }
    }

@app.get("/api/auth/me")
async def get_current_user(admin: Dict = Depends(get_current_admin_from_token)):
    """Отримати інформацію про поточного адміна"""
    return {
        "id": admin["id"],
        "username": admin["username"],
        "email": admin["email"],
        "first_name": admin["first_name"],
        "last_name": admin["last_name"],
        "is_superadmin": admin["is_superadmin"],
        "last_login_at": admin["last_login_at"]
    }

@app.get("/api/admins")
async def get_admins(admin: Dict = Depends(get_current_admin_from_token)):
    """Отримати список адмінів"""
    if not check_admin_permission(admin, "manage_admins"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    db = get_database()
    cursor = db.cursor(dictionary=True)
    
    try:
        cursor.execute("""
            SELECT id, username, email, first_name, last_name, role, 
                   is_active, is_superadmin, can_manage_users, can_manage_payments,
                   can_manage_settings, can_manage_admins, created_at, last_login_at
            FROM admins 
            ORDER BY created_at DESC
        """)
        
        admins = cursor.fetchall() or []
        
        # Форматуємо дати
        for admin_item in admins:
            if admin_item.get('created_at'):
                admin_item['created_at'] = admin_item['created_at'].isoformat()
            if admin_item.get('last_login_at'):
                admin_item['last_login_at'] = admin_item['last_login_at'].isoformat()
        
        return {"admins": admins}
        
    finally:
        cursor.close()
        db.close()

@app.post("/api/admins")
async def create_admin(admin_data: AdminCreate, admin: Dict = Depends(get_current_admin_from_token)):
    """Створити нового адміна"""
    if not check_admin_permission(admin, "manage_admins"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    db = get_database()
    cursor = db.cursor()
    
    try:
        # Перевіряємо чи не існує вже адмін з таким username або email
        cursor.execute(
            "SELECT id FROM admins WHERE username = %s OR email = %s",
            (admin_data.username, admin_data.email)
        )
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail="Username or email already exists")
        
        # Хешуємо пароль
        password_hash = hash_password(admin_data.password)
        
        # Створюємо адміна
        cursor.execute("""
            INSERT INTO admins (
                username, email, password_hash, first_name, last_name, is_active
            ) VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            admin_data.username,
            admin_data.email,
            password_hash,
            admin_data.first_name,
            admin_data.last_name,
            admin_data.is_active
        ))
        
        db.commit()
        admin_id = cursor.lastrowid
        
        return {"success": True, "message": "Admin created successfully", "id": admin_id}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating admin: {str(e)}")
    finally:
        cursor.close()
        db.close()

@app.put("/api/admins/{admin_id}")
async def update_admin(admin_id: int, admin_data: AdminUpdate, admin: Dict = Depends(get_current_admin_from_token)):
    """Оновити адміна"""
    if not check_admin_permission(admin, "manage_admins"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    db = get_database()
    cursor = db.cursor()
    
    try:
        # Перевіряємо чи існує адмін
        cursor.execute("SELECT id FROM admins WHERE id = %s", (admin_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="Admin not found")
        
        # Формуємо запит на оновлення
        update_fields = []
        update_values = []
        
        for field, value in admin_data.dict(exclude_none=True).items():
            update_fields.append(f"{field} = %s")
            update_values.append(value)
        
        if update_fields:
            update_values.append(admin_id)
            cursor.execute(
                f"UPDATE admins SET {', '.join(update_fields)}, updated_at = NOW() WHERE id = %s",
                update_values
            )
            db.commit()
        
        return {"success": True, "message": "Admin updated successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating admin: {str(e)}")
    finally:
        cursor.close()
        db.close()

@app.delete("/api/admins/{admin_id}")
async def delete_admin(admin_id: int, admin: Dict = Depends(get_current_admin_from_token)):
    """Видалити адміна"""
    if not check_admin_permission(admin, "manage_admins"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Не можна видалити самого себе
    if admin["id"] == admin_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    db = get_database()
    cursor = db.cursor()
    
    try:
        cursor.execute("DELETE FROM admins WHERE id = %s", (admin_id,))
        
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Admin not found")
        
        db.commit()
        return {"success": True, "message": "Admin deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting admin: {str(e)}")
    finally:
        cursor.close()
        db.close()

@app.post("/api/admins/{admin_id}/change-password")
async def change_admin_password(admin_id: int, password_data: PasswordChange, admin: Dict = Depends(get_current_admin_from_token)):
    """Змінити пароль адміна"""
    # Можна змінити свій пароль або якщо є права manage_admins
    if admin["id"] != admin_id and not check_admin_permission(admin, "manage_admins"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    db = get_database()
    cursor = db.cursor(dictionary=True)
    
    try:
        # Отримуємо поточний пароль
        cursor.execute("SELECT password_hash FROM admins WHERE id = %s", (admin_id,))
        admin_data = cursor.fetchone()
        
        if not admin_data:
            raise HTTPException(status_code=404, detail="Admin not found")
        
        # Перевіряємо поточний пароль (тільки якщо змінюємо свій)
        if admin["id"] == admin_id:
            if not verify_password(password_data.current_password, admin_data["password_hash"]):
                raise HTTPException(status_code=400, detail="Current password is incorrect")
        
        # Хешуємо новий пароль
        new_password_hash = hash_password(password_data.new_password)
        
        # Оновлюємо пароль
        cursor.execute(
            "UPDATE admins SET password_hash = %s, updated_at = NOW() WHERE id = %s",
            (new_password_hash, admin_id)
        )
        
        db.commit()
        return {"success": True, "message": "Password changed successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error changing password: {str(e)}")
    finally:
        cursor.close()
        db.close()

@app.get("/api/settings")
async def get_settings(admin: Dict = Depends(get_current_admin_flexible)):
    """Отримати налаштування системи"""
    # Для адмін панелі повертаємо всі налаштування, включаючи sensitive
    all_settings = settings_manager.get_all_settings(include_sensitive=True)
    
    return {
        "bot_token": all_settings.get("bot_token", ""),
        "webhook_url": all_settings.get("webhook_url", ""),
        "stripe_publishable_key": all_settings.get("stripe_publishable_key", ""),
        "stripe_secret_key": all_settings.get("stripe_secret_key", ""),
        "stripe_webhook_secret": all_settings.get("stripe_webhook_secret", ""),
        "subscription_price": all_settings.get("subscription_price", 15.0),
        "subscription_currency": all_settings.get("subscription_currency", "eur"),
        "database_status": "Connected",
        "bot_status": "Running",
        "webhook_status": "Active",
        "app_name": all_settings.get("app_name", "Upgrade Studio Bot"),
        "support_email": all_settings.get("support_email", "support@upgradestudio.com"),
        "maintenance_mode": all_settings.get("maintenance_mode", False)
    }

@app.get("/api/settings/all")
async def get_all_settings(admin: Dict = Depends(get_current_admin_from_token)):
    """Отримати всі налаштування для редагування"""
    if not check_admin_permission(admin, "manage_settings"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    db = get_database()
    cursor = db.cursor(dictionary=True)
    
    try:
        cursor.execute("""
            SELECT s.*, a.username as updated_by_username
            FROM system_settings s
            LEFT JOIN admins a ON s.updated_by = a.id
            ORDER BY s.category, s.key
        """)
        
        settings_list = cursor.fetchall() or []
        
        # Дешифруємо значення і форматуємо дати
        from database.encryption import decrypt_setting
        
        for setting in settings_list:
            if not setting["is_sensitive"]:
                setting["decrypted_value"] = decrypt_setting(setting["encrypted_value"], setting["value_type"])
            else:
                setting["decrypted_value"] = "***HIDDEN***"
            
            if setting.get('created_at'):
                setting['created_at'] = setting['created_at'].isoformat()
            if setting.get('updated_at'):
                setting['updated_at'] = setting['updated_at'].isoformat()
        
        return {"settings": settings_list}
        
    finally:
        cursor.close()
        db.close()

@app.put("/api/settings/{setting_key}")
async def update_setting(setting_key: str, setting_data: SettingUpdate, admin: Dict = Depends(get_current_admin_flexible)):
    """Оновити налаштування"""
    if not check_admin_permission(admin, "manage_settings"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    try:
        success = settings_manager.set(
            key=setting_key,
            value=setting_data.value,
            value_type=setting_data.value_type,
            category=setting_data.category,
            is_sensitive=setting_data.is_sensitive,
            description=setting_data.description,
            updated_by=admin["id"]
        )
        
        if success:
            # Очищаємо кеш налаштувань для динамічного оновлення
            settings.invalidate_cache()
            return {"success": True, "message": "Setting updated successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to update setting")
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating setting: {str(e)}")

class PaymentSuccessRequest(BaseModel):
    telegram_id: int

@app.post("/trigger_payment_success")
async def trigger_payment_success(request: PaymentSuccessRequest):
    """Викликає обробник успішної оплати в боті"""
    try:
        # Імпортуємо bot_instance динамічно
        from main import bot_instance
        
        # Викликаємо обробник успішної оплати
        import asyncio
        asyncio.create_task(bot_instance.handle_successful_payment(request.telegram_id))
        
        return {"success": True, "message": f"Payment success handler triggered for user {request.telegram_id}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error triggering payment success: {str(e)}")

# ============================================
# BROADCASTS ENDPOINTS
# ============================================

class BroadcastCreate(BaseModel):
    target_group: str  # 'active', 'inactive', 'no_payment'
    title: Optional[str] = None  # Заголовок розсилки
    message_text: Optional[str] = None
    attachment_type: Optional[str] = None  # 'image', 'file', 'link'
    attachment_url: Optional[str] = None
    button_text: Optional[str] = None
    button_url: Optional[str] = None
    message_blocks: Optional[list] = None  # Всі блоки повідомлення

@app.get("/api/broadcasts")
async def get_broadcasts(
    page: int = 1,
    limit: int = 50,
    admin: Dict = Depends(get_current_admin_flexible)
) -> Dict[str, Any]:
    """Отримати список розсилок"""
    try:
        db = get_database()
        cursor = db.cursor(dictionary=True)
        
        offset = (page - 1) * limit
        
        # Отримуємо розсилки з інформацією про адміна
        cursor.execute("""
            SELECT 
                b.id,
                b.target_group,
                b.title,
                b.message_text,
                b.attachment_type,
                b.attachment_url,
                b.button_text,
                b.button_url,
                b.status,
                b.total_recipients,
                b.sent_count,
                b.failed_count,
                b.created_at,
                b.started_at,
                b.completed_at,
                b.error_log,
                b.full_log,
                a.username as created_by_username
            FROM broadcasts b
            LEFT JOIN admins a ON b.created_by = a.id
            ORDER BY b.created_at DESC
            LIMIT %s OFFSET %s
        """, (limit, offset))
        
        broadcasts = cursor.fetchall()
        
        # Total count
        cursor.execute("SELECT COUNT(*) as total FROM broadcasts")
        result = cursor.fetchone()
        total = result["total"] if result else 0
        
        cursor.close()
        db.close()
        
        return {
            "broadcasts": broadcasts,
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching broadcasts: {str(e)}")

@app.get("/api/broadcasts/stats")
async def get_broadcast_stats(admin: Dict = Depends(get_current_admin_flexible)) -> Dict[str, Any]:
    """Отримати статистику по групам користувачів для розсилок"""
    try:
        db = get_database()
        cursor = db.cursor(dictionary=True)
        
        # Активні підписники
        cursor.execute("SELECT COUNT(*) as count FROM users WHERE subscription_active = 1")
        result = cursor.fetchone()
        active_count = result["count"] if result else 0
        
        # Неактивні (були активні, але більше ні)
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM users 
            WHERE subscription_active = 0 
            AND stripe_subscription_id IS NOT NULL
        """)
        result = cursor.fetchone()
        inactive_count = result["count"] if result else 0
        
        # Не оплатили (запустили бота, але немає subscription_id)
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM users 
            WHERE stripe_subscription_id IS NULL
        """)
        result = cursor.fetchone()
        no_payment_count = result["count"] if result else 0
        
        cursor.close()
        db.close()
        
        return {
            "active": active_count,
            "inactive": inactive_count,
            "no_payment": no_payment_count
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching broadcast stats: {str(e)}")

@app.post("/api/broadcasts/upload")
async def upload_broadcast_file(
    file: UploadFile = File(...),
    admin: Dict = Depends(get_current_admin_flexible)
) -> Dict[str, Any]:
    """Завантажити файл для розсилки"""
    if not check_admin_permission(admin, "manage_broadcasts"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    try:
        # Перевіряємо тип файлу
        allowed_types = {
            'image/jpeg', 'image/png', 'image/gif', 'image/webp',
            'video/mp4', 'video/mpeg', 'video/quicktime',
            'application/pdf', 'application/zip',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        if file.content_type not in allowed_types:
            raise HTTPException(
                status_code=400,
                detail=f"File type {file.content_type} not allowed. Allowed types: images, videos, PDF, ZIP, DOCX, XLSX"
            )
        
        # Перевіряємо розмір (максимум 50MB)
        max_size = 50 * 1024 * 1024  # 50MB
        file_content = await file.read()
        if len(file_content) > max_size:
            raise HTTPException(status_code=400, detail="File too large. Maximum size is 50MB")
        
        # Генеруємо унікальне ім'я файлу
        file_extension = file.filename.split('.')[-1] if '.' in file.filename else ''
        unique_filename = f"{uuid.uuid4()}.{file_extension}"
        file_path = BROADCASTS_DIR / unique_filename
        
        # Обробка зображень - ресайз якщо занадто великі
        if file.content_type.startswith('image/'):
            try:
                img = Image.open(io.BytesIO(file_content))
                
                # Telegram обмеження: рекомендовано до 5000x5000, макс 10000x10000
                max_dimension = 4096
                if img.width > max_dimension or img.height > max_dimension:
                    # Зберігаємо пропорції
                    ratio = min(max_dimension / img.width, max_dimension / img.height)
                    new_size = (int(img.width * ratio), int(img.height * ratio))
                    
                    # Ресайз з високою якістю
                    img = img.resize(new_size, Image.Resampling.LANCZOS)
                    
                    # Конвертуємо в RGB якщо потрібно (для JPEG)
                    if img.mode in ('RGBA', 'LA', 'P'):
                        rgb_img = Image.new('RGB', img.size, (255, 255, 255))
                        if img.mode == 'P':
                            img = img.convert('RGBA')
                        rgb_img.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                        img = rgb_img
                    
                    # Зберігаємо оптимізоване зображення
                    img.save(file_path, format='JPEG', quality=85, optimize=True)
                else:
                    # Зберігаємо оригінал якщо розмір прийнятний
                    with open(file_path, 'wb') as f:
                        f.write(file_content)
            except Exception as e:
                # Якщо помилка обробки - зберігаємо оригінал
                with open(file_path, 'wb') as f:
                    f.write(file_content)
        else:
            # Зберігаємо файл як є (відео, документи)
            with open(file_path, 'wb') as f:
                f.write(file_content)
        
        # Визначаємо тип вкладення
        attachment_type = None
        if file.content_type.startswith('image/'):
            attachment_type = 'image'
        elif file.content_type.startswith('video/'):
            attachment_type = 'video'
        else:
            attachment_type = 'document'  # Змінено з 'file' на 'document' для сумісності з фронтендом
        
        # Формуємо URL для доступу до файлу
        file_url = f"/uploads/broadcasts/{unique_filename}"
        
        return {
            "success": True,
            "filename": file.filename,
            "url": file_url,
            "attachment_type": attachment_type,
            "size": len(file_content),
            "content_type": file.content_type
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error uploading file: {str(e)}")

@app.delete("/api/broadcasts/file")
async def delete_broadcast_file(
    file_url: str,
    admin: Dict = Depends(get_current_admin_flexible)
) -> Dict[str, Any]:
    """Видалити файл розсилки"""
    if not check_admin_permission(admin, "manage_broadcasts"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    try:
        # Отримуємо шлях до файлу з URL
        if not file_url.startswith('/uploads/broadcasts/'):
            raise HTTPException(status_code=400, detail="Invalid file URL")
        
        filename = file_url.split('/')[-1]
        file_path = BROADCASTS_DIR / filename
        
        # Видаляємо файл якщо він існує
        if file_path.exists():
            os.remove(file_path)
            return {"success": True, "message": "File deleted"}
        else:
            return {"success": True, "message": "File not found"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting file: {str(e)}")

@app.post("/api/broadcasts")
async def create_broadcast(
    broadcast_data: BroadcastCreate,
    admin: Dict = Depends(get_current_admin_flexible)
) -> Dict[str, Any]:
    """Створити нову розсилку"""
    if not check_admin_permission(admin, "manage_broadcasts"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    try:
        from database.models import DatabaseManager, Broadcast, BroadcastQueue, User
        
        # Логування отриманих даних
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Creating broadcast with data:")
        logger.info(f"  target_group: {broadcast_data.target_group}")
        logger.info(f"  title: '{broadcast_data.title}'")
        logger.info(f"  message_text: {broadcast_data.message_text[:50] if broadcast_data.message_text else None}")
        logger.info(f"  attachment_type: {broadcast_data.attachment_type}")
        
        with DatabaseManager() as db:
            # Створюємо розсилку
            broadcast = Broadcast(
                created_by=admin["id"],
                target_group=broadcast_data.target_group,
                title=broadcast_data.title,
                message_text=broadcast_data.message_text,
                attachment_type=broadcast_data.attachment_type,
                attachment_url=broadcast_data.attachment_url,
                button_text=broadcast_data.button_text,
                button_url=broadcast_data.button_url,
                status='pending'
            )
            
            # Зберігаємо message_blocks якщо передані
            if broadcast_data.message_blocks:
                import json
                broadcast.message_blocks = json.dumps(broadcast_data.message_blocks)
            
            db.add(broadcast)
            db.flush()  # Отримуємо ID
            
            logger.info(f"Broadcast created with ID: {broadcast.id}, title in object: '{broadcast.title}'")
            
            # Вибираємо користувачів за групою
            if broadcast_data.target_group == 'active':
                users = db.query(User).filter(User.subscription_active == True).all()
            elif broadcast_data.target_group == 'inactive':
                users = db.query(User).filter(
                    User.subscription_active == False,
                    User.stripe_subscription_id.isnot(None)
                ).all()
            elif broadcast_data.target_group == 'no_payment':
                users = db.query(User).filter(User.stripe_subscription_id.is_(None)).all()
            else:
                raise HTTPException(status_code=400, detail="Invalid target group")
            
            # Додаємо користувачів в чергу
            for user in users:
                queue_item = BroadcastQueue(
                    broadcast_id=broadcast.id,
                    user_id=user.id,
                    telegram_id=user.telegram_id
                )
                db.add(queue_item)
            
            broadcast.total_recipients = len(users)
            db.commit()
            
            return {
                "success": True,
                "broadcast_id": broadcast.id,
                "total_recipients": len(users)
            }
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating broadcast: {str(e)}")

@app.get("/api/broadcasts/{broadcast_id}")
async def get_broadcast_detail(
    broadcast_id: int,
    admin: Dict = Depends(get_current_admin_flexible)
) -> Dict[str, Any]:
    """Отримати детальну інформацію про розсилку"""
    try:
        db = get_database()
        cursor = db.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                b.*,
                a.username as created_by_username
            FROM broadcasts b
            LEFT JOIN admins a ON b.created_by = a.id
            WHERE b.id = %s
        """, (broadcast_id,))
        
        broadcast = cursor.fetchone()
        
        if not broadcast:
            raise HTTPException(status_code=404, detail="Broadcast not found")
        
        cursor.close()
        db.close()
        
        return {"broadcast": broadcast}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching broadcast: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)